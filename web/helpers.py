from openpyxl import load_workbook
from .models import Cliente, Calle
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from sigabd import sigabdConnector
from django.http import JsonResponse
from datetime import datetime
import json
import logging

logger = logging.getLogger(__name__)

def load_clientes():

    wb = load_workbook('clientes_sin_localizar.xlsx')
    ws = wb.get_sheet_by_name('clientes')
    clientes = []
    for r in range(2, ws.max_row + 1):
        if ws['B%s' % (r)].value is None:
            ws['B%s' % (r)].value = ""
        if ws['C%s' % (r)].value is None:
            ws['C%s' % (r)].value = ""
        if ws['D%s' % (r)].value is None:
            ws['D%s' % (r)].value = ""
        if ws['E%s' % (r)].value is None and ws['G%s' % (r)].value is None:
            continue
        if ws['E%s' % (r)].value is None:
            ws['E%s' % (r)].value = ""
        if ws['G%s' % (r)].value is None:
            ws['G%s' % (r)].value = ""
        if ws['H%s' % (r)].value is None:
            ws['H%s' % (r)].value = ""
        if ws['I%s' % (r)].value is None:
            ws['I%s' % (r)].value = ""
        if ws['J%s' % (r)].value == "NO":
            geocode = False
        else:
            geocode = True
        clientes.append(
            Cliente(
                clientenro=ws['A%s' % (r)].value,
                nombre=ws['B%s' % (r)].value + " " + ws['C%s' % (r)].value,
                direccion=ws['D%s' % (r)].value + " " +
                str(ws['E%s' % (r)].value),
                tira=ws['G%s' % (r)].value,
                piso=ws['H%s' % (r)].value,
                depto=ws['I%s' % (r)].value,
                geocode=geocode
            )
        )
    Cliente.objects.bulk_create(clientes)


def load_calles():
    wb = load_workbook('calles_completo.xlsx')
    ws = wb['calles_completo']
    for r in range(2, ws.max_row + 1):
        if ws['F%s' % (r)].value is None or ws['F%s' % (r)].value == "":
            continue
        else:
            calleidsiga = ws['F%s' % (r)].value
        nombre = ws['A%s' % (r)].value
        if ws['B%s' % (r)].value is None or ws['A%s' % (r)].value == "":
            lim_inf = None
        else:
            lim_inf = ws['B%s' % (r)].value
        if ws['C%s' % (r)].value is None or ws['C%s' % (r)].value == "":
            lim_sup = None
        else:
            lim_sup = ws['C%s' % (r)].value
        if ws['D%s' % (r)].value == 'SI':
            geocode = True
        else:
            geocode = False
        if ws['E%s' % (r)].value == 'SI':
            osm_geocode = True
        elif ws['E%s' % (r)].value == 'NO':
            osm_geocode = False
        else:
            osm_geocode = False
        if ws['G%s' % (r)].value == 'SI':
            calle_chica = True
        else:
            calle_chica = False
        try:
            c = Calle.objects.get(calleidsiga=calleidsiga)
        except (ObjectDoesNotExist, MultipleObjectsReturned):
            continue
        c.nombre = nombre
        c.limite_inferior = lim_inf
        c.limite_superior = lim_sup
        c.calleidsiga = calleidsiga
        c.geocode = geocode
        c.osm_geocode = osm_geocode
        c.calle_chica = calle_chica
        c.save()
    return True


def update_estado():
    clientes = Cliente.objects.filter(estado__isnull=True)
    for c in clientes:
        con = sigabdConnector("SIGAARG", "SIGAARG")
        r = con.get_estado_for_cliente(c.clientenro)
        c.estado = r
        c.save()
    return True


def compare_to_gis():
    with open('ubicados.geojson') as infile:
        f = json.load(infile)
    cont = 0
    for feature in f["features"]:
        cliente = Cliente.objects.filter(
            clientenro=feature["properties"]["clientenro"],
            latitud_4326__isnull=True
        )
        try:
            cli = cliente[0]
            cli.latitud_221712 = feature["geometry"]["coordinates"][0]
            cli.longitud_221712 = feature["geometry"]["coordinates"][1]
            cli.save()
            cont += 1
        except IndexError:
            continue
    return JsonResponse(json.dumps(cont))


def load_tiras_coords():
    wb = load_workbook('barrios_ushuaia.xlsx')
    ws = wb.get_sheet_by_name('tiras')
    """640viv
    """
    r = 1
    calleidsiga = ws['E%s' % (r)].value
    cant = 0
    for r in range(2, 49):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    """64 viviendas
    """
    r = 49
    calleidsiga = ws['E%s' % (r)].value
    for r in range(50, 54):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    """monte gallinero

    """ 
    r = 54
    calleidsiga = ws['E%s' % (r)].value
    for r in range(55, 96):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    """30 viviendas
    """
    r = 96
    calleidsiga = ws['E%s' % (r)].value
    for r in range(97, 107):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    """245 VIVIENDAS
    """
    r = 107
    calleidsiga = ws['E%s' % (r)].value
    for r in range(108, 139):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    """60 massiotra
    """
    r = 139
    calleidsiga = ws['E%s' % (r)].value
    for r in range(140, 160):
        clientes = Cliente.objects.filter(
            latitud_4326__isnull=True,
            calle__calleidsiga=calleidsiga,
            tira=ws['B%s' % (r)].value
        )
        for c in clientes:
            c.latitud_4326 = ws['C%s' % (r)].value
            c.longitud_4326 = ws['D%s' % (r)].value
            c.precision = 0
            c.fecha_posicion = datetime.now()
            c.save()
            cant += 1
    return JsonResponse({"cantidad": cant})
